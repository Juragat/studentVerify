from flask import Flask, request, jsonify
import random
import string
from openpyxl import load_workbook
from flask_cors import CORS
CORS(app)

# Initialize Flask app
app = Flask(__name__)

# Store one-time codes temporarily
codes = {}

# Teacher Login Route
@app.route('/teacher/login', methods=['POST'])
def teacher_login():
    data = request.json
    username = data['username']
    password = data['password']
    # Replace with proper validation (e.g., from a database)
    if username == "teacher1" and password == "securepassword":  
        return jsonify({"message": "Login successful"}), 200
    return jsonify({"message": "Invalid credentials"}), 401

# Generate One-Time Code Route
@app.route('/teacher/generate_code', methods=['POST'])
def generate_code():
    code = ''.join(random.choices(string.ascii_uppercase + string.digits, k=6))
    codes[code] = True  # Store the code as valid
    return jsonify({"code": code}), 200

# Mark Attendance Route
@app.route('/student/mark', methods=['POST'])
def mark_attendance():
    data = request.json
    student_id = data['studentId']
    code = data['code']

    # Check if the code is valid
    if code not in codes or not codes[code]:
        return jsonify({"message": "Invalid or expired code"}), 400

    # Update attendance in Excel
    try:
        workbook = load_workbook('attendance.xlsx')
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2):  # Assuming first row is headers
            if row[0].value == student_id:
                row[2].value = "Present"  # Mark attendance
                workbook.save('attendance.xlsx')
                return jsonify({"message": "Attendance marked successfully"}), 200
    except Exception as e:
        return jsonify({"message": f"Error updating attendance: {str(e)}"}), 500

    return jsonify({"message": "Student ID not found"}), 404

# Run the Flask app (for local testing)
if __name__ == '__main__':
    app.run(debug=True)
