from openpyxl import load_workbook

@app.route('/student/mark', methods=['POST'])
def mark_attendance():
    data = request.json
    student_id = data['studentId']
    code = data['code']

    # Check if code is valid
    if code not in codes or not codes[code]:
        return jsonify({"message": "Invalid or expired code"}), 400

    # Update attendance in Excel
    workbook = load_workbook('attendance.xlsx')
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2):  # Assuming first row is headers
        if row[0].value == student_id:
            row[2].value = "Present"  # Mark attendance
            workbook.save('attendance.xlsx')
            return jsonify({"message": "Attendance marked successfully"}), 200

    return jsonify({"message": "Student ID not found"}), 404
